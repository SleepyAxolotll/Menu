from fastapi import FastAPI, UploadFile
from fastapi.responses import JSONResponse, FileResponse
from bs4 import BeautifulSoup
from openpyxl import Workbook
import tempfile
import os
import uvicorn

app = FastAPI()

@app.get("/")
def read_root():
    return {"message": "Welcome to the Menu Processing App!"}
    
@app.post("/process_menu")
async def process_menu(file_upload: UploadFile):
    try:
        # Create a temporary directory to store the Excel file
        with tempfile.TemporaryDirectory() as temp_dir:
            excel_file_path = os.path.join(temp_dir, "menu.xlsx")
            wb = Workbook()
            ws = wb.active

            html_content = await file_upload.read()
            soup = BeautifulSoup(html_content, 'html.parser')

            # Extract data from HTML and populate the Excel workbook
            date_wrappers = soup.find_all(class_="day-name")
            meal_names = [meal_name.h3.get_text(strip=True) for meal_name in soup.find_all(class_="meal-name")]
            ul_elements = soup.find_all('ul')

            food_times_list = [x.getText(strip=True) for x in date_wrappers]
            for index, value in enumerate(food_times_list, start=2):
                ws.cell(row=index, column=1, value=value)

            meals_list = ["BREAKFAST", "LUNCH", "DINNER"]
            for col, meal in enumerate(meals_list, start=2):
                ws.cell(row=1, column=col, value=meal)

            day_tracker = 2
            meal_name_counter = 0

            for ul in ul_elements:
                current_meal_name = meal_names[meal_name_counter]
                meal_name_counter += 1

                food_elements_list = [food_item.find('div').getText(strip=True) for food_item in ul.find_all('li', class_='food')]
                list_string = ', '.join(map(str, food_elements_list))

                if "BREAKFAST" in current_meal_name:
                    ws.cell(row=day_tracker, column=2, value=list_string)
                elif "LUNCH" in current_meal_name:
                    ws.cell(row=day_tracker, column=3, value=list_string)
                else:
                    ws.cell(row=day_tracker, column=4, value=list_string)
                    day_tracker += 1

            # Save the workbook to the temporary file
            wb.save(excel_file_path)

            # Return the Excel file as a downloadable response
            return FileResponse(excel_file_path, filename="menu.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        # Handle and log specific exceptions for debugging
        return JSONResponse(status_code=500, content={"message": f"An error occurred: {str(e)}"})

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
